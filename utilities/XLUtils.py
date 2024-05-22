import openpyxl
from openpyxl.styles import PatternFill

def Rows(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_row)

def Column(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_column)

def Readdata(file,sheetName,rowNo,columnNo):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(rowNo,columnNo).value

def Writedata(file,sheetName,rowNo,columnNo,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(rowNo,columnNo).value=data
    workbook.save(file)

def fillGreencolor(file,sheetName,rowNo,columnNo):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    greenFill=PatternFill(start_color='60b212',
                          end_color='60b212',
                          fill_type='solid')
    sheet.cell(rowNo,columnNo).fill=greenFill
    workbook.save(file)

def fillRedcolor(file,sheetName,rowNo,columnNo):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    redFill=PatternFill(start_color='ff0000',
                        end_color='ff0000',
                        fill_type='solid')
    sheet.cell(rowNo,columnNo).fill=redFill
    workbook.save(file)

